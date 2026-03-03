import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

def build_orders_xlsx_with_images(rows: list[dict], images_dir: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "buyurtmalar"

    headers = [
        "ID raqami", "Yaratilgan vaqt", "Hudud", "Manzil", "Kg",
        "Qabul qiluvchi F.I.O", "Qabul qiluvchi tel", "Google Maps",
        "To‘lov turi", "To‘lov holati", "Status",
        "Jami (TRY)", "Jami (UZS)",
        "Pasport old rasmi", "Pasport orqa rasmi", "Yuk ro‘yxati rasmi", "To‘lov cheki"
    ]
    ws.append(headers)

    widths = [18, 19, 14, 44, 6, 22, 18, 34, 12, 18, 12, 14, 16, 20, 20, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row_idx = 2

    def put(col: int, filename: str):
        path = os.path.join(images_dir, filename)
        if not os.path.exists(path):
            return
        img = XLImage(path)
        img.width = 160
        img.height = 90
        ws.row_dimensions[row_idx].height = 90
        ws.add_image(img, f"{get_column_letter(col)}{row_idx}")

    for r in rows:
        ws.append([
            r.get("order_code"), r.get("created_at"), r.get("district"), r.get("address"), r.get("kg"),
            r.get("receiver_name"), r.get("receiver_phone"), r.get("maps_link"),
            r.get("payment_type"), r.get("payment_status"), r.get("status"),
            r.get("total_try"), r.get("total_uzs"),
            "", "", "", ""
        ])
        oc = r.get("order_code") or str(r.get("id"))
        put(14, f"{oc}_passport_front.jpg")
        put(15, f"{oc}_passport_back.jpg")
        put(16, f"{oc}_items.jpg")
        put(17, f"{oc}_payment.jpg")
        row_idx += 1

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio